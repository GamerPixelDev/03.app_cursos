"""Utilidades para exportar datos a hojas de cálculo Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models import alumnos, cursos, matriculas
from models.export_paths import resolver_ruta_exportacion


def _autoajustar_columnas(columnas: Iterable[Iterable]) -> None:
    """Ajusta el ancho de todas las columnas en función de su contenido."""

    for columna in columnas:
        max_len = max(len(str(cell.value or "")) for cell in columna)
        columna[0].parent.column_dimensions[columna[0].column_letter].width = max_len + 2


def _guardar_libro(libro: Workbook, ruta_archivo: Path) -> str:
    ruta_archivo.parent.mkdir(parents=True, exist_ok=True)
    libro.save(ruta_archivo)
    return str(ruta_archivo)


def _destino(nombre_archivo: str, *, usuario: str | None, ruta_destino: str | None) -> Path:
    return resolver_ruta_exportacion(usuario=usuario, ruta_destino=ruta_destino) / nombre_archivo


def exportar_alumnos_excel(*, usuario: str | None = None, ruta_destino: str | None = None) -> str:
    """Exporta el listado de alumnos a Excel.

    Parameters
    ----------
    usuario: str | None
        Usuario que ejecuta la exportación. Si tiene configurada una ruta
        personalizada se empleará automáticamente.
    ruta_destino: str | None
        Ruta manual proporcionada en tiempo de ejecución.
    """

    datos = alumnos.obtener_alumnos()
    archivo = _destino("alumnos.xlsx", usuario=usuario, ruta_destino=ruta_destino)

    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    columnas = [
        "NIF", "Nombre", "Apellidos", "Localidad", "Código Postal",
        "Teléfono", "Email", "Sexo", "Edad", "Estudios", "Estado Laboral"
    ]
    ws.append(columnas)

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center")
        celda.fill = PatternFill("solid", fgColor="3E64FF")

    for fila in datos:
        ws.append(fila)

    _autoajustar_columnas(ws.columns)
    return _guardar_libro(wb, archivo)


def exportar_cursos_excel(*, usuario: str | None = None, ruta_destino: str | None = None) -> str:
    datos = cursos.obtener_cursos()
    if not datos:
        raise ValueError("No hay cursos para exportar.")

    archivo = _destino("cursos.xlsx", usuario=usuario, ruta_destino=ruta_destino)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cursos"
    columnas = [
        "Código", "Nombre", "Fecha Inicio", "Fecha Fin",
        "Lugar", "Modalidad", "Horas", "Responsable"
    ]
    ws.append(columnas)

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center")
        celda.fill = PatternFill("solid", fgColor="3E64FF")

    for fila in datos:
        ws.append(fila)

    _autoajustar_columnas(ws.columns)
    return _guardar_libro(wb, archivo)


def exportar_matriculas_excel(*, usuario: str | None = None, ruta_destino: str | None = None) -> str:
    datos = matriculas.obtener_matriculas()
    if not datos:
        raise ValueError("No hay matrículas para exportar.")

    archivo = _destino("matriculas.xlsx", usuario=usuario, ruta_destino=ruta_destino)

    wb = Workbook()
    ws = wb.active
    ws.title = "Matrículas"
    columnas = [
        "NIF", "Nombre Alumno", "Código Curso", "Curso", "Fecha Insc."
    ]
    ws.append(columnas)

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center")
        celda.fill = PatternFill("solid", fgColor="3E64FF")

    for fila in datos:
        ws.append(fila)

    _autoajustar_columnas(ws.columns)
    return _guardar_libro(wb, archivo)