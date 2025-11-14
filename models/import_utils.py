"""Herramientas de importación con validaciones para proteger los datos."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Iterable, Optional

from openpyxl import load_workbook

from models import alumnos, cursos, matriculas


@dataclass
class ResultadoImportacion:
    entidad: str
    nuevos: int = 0
    duplicados: int = 0
    errores: int = 0

    def como_dict(self) -> dict[str, int | str]:
        return {
            "entidad": self.entidad,
            "nuevos": self.nuevos,
            "duplicados": self.duplicados,
            "errores": self.errores,
        }


def _a_cadena(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
    return str(valor).strip()


def _a_entero(valor) -> Optional[int]:
    if valor in (None, ""):
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def _normalizar_fecha(valor) -> Optional[str]:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    cadena = _a_cadena(valor)
    if not cadena:
        return None
    formatos = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")
    for fmt in formatos:
        try:
            return datetime.strptime(cadena, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _email_valido(email: str) -> bool:
    if not email:
        return True  # Campo opcional.
    return "@" in email and "." in email.split("@")[-1]


def _fila_vacia(fila: Iterable) -> bool:
    return all(celda in (None, "") for celda in fila)


def importar_alumnos_desde_excel(ruta: str) -> dict[str, int | str]:
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]  # saltar cabecera

    resumen = ResultadoImportacion("alumnos")

    for fila in filas:
        if not fila or _fila_vacia(fila):
            continue

        datos = list(fila) + [None] * max(0, 11 - len(fila))
        nif = _a_cadena(datos[0]).upper()
        nombre = _a_cadena(datos[1])
        apellidos = _a_cadena(datos[2])
        if not (nif and nombre and apellidos):
            resumen.errores += 1
            continue

        codigo_postal = _a_cadena(datos[4])
        telefono = _a_cadena(datos[5])
        email = _a_cadena(datos[6]).lower()
        if not _email_valido(email):
            resumen.errores += 1
            continue

        sexo = _a_cadena(datos[7]).upper() or None
        if sexo and sexo not in {"M", "F", "H", "HOMBRE", "MUJER"}:
            sexo = sexo[:1]

        edad = _a_entero(datos[8])
        estudios = _a_cadena(datos[9])
        estado_laboral = _a_cadena(datos[10])

        ok = alumnos.crear_alumno(
            nif,
            nombre,
            apellidos,
            _a_cadena(datos[3]),  # localidad
            codigo_postal,
            telefono,
            email,
            sexo,
            edad,
            estudios,
            estado_laboral,
        )
        if ok:
            resumen.nuevos += 1
        else:
            resumen.duplicados += 1

    return resumen.como_dict()


def importar_cursos_desde_excel(ruta: str) -> dict[str, int | str]:
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]

    resumen = ResultadoImportacion("cursos")

    for fila in filas:
        if not fila or _fila_vacia(fila):
            continue

        datos = list(fila) + [None] * max(0, 8 - len(fila))
        codigo = _a_cadena(datos[0]).upper()
        nombre = _a_cadena(datos[1])
        fecha_inicio = _normalizar_fecha(datos[2])
        fecha_fin = _normalizar_fecha(datos[3])
        horas = _a_entero(datos[6])

        if not (codigo and nombre and fecha_inicio and fecha_fin and horas is not None):
            resumen.errores += 1
            continue

        ok = cursos.crear_curso(
            codigo,
            nombre,
            fecha_inicio,
            fecha_fin,
            _a_cadena(datos[4]),
            _a_cadena(datos[5]),
            horas,
            _a_cadena(datos[7]),
        )

        if ok:
            resumen.nuevos += 1
        else:
            resumen.duplicados += 1

    return resumen.como_dict()


def importar_matriculas_desde_excel(ruta: str) -> dict[str, int | str]:
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]

    resumen = ResultadoImportacion("matrículas")
    cache_alumnos: dict[str, bool] = {}
    cache_cursos: dict[str, bool] = {}

    for fila in filas:
        if not fila or _fila_vacia(fila):
            continue

        datos = list(fila) + [None] * max(0, 5 - len(fila))
        nif = _a_cadena(datos[0]).upper()
        codigo = _a_cadena(datos[2]).upper()
        fecha = _normalizar_fecha(datos[4])

        if not (nif and codigo):
            resumen.errores += 1
            continue

        if nif not in cache_alumnos:
            cache_alumnos[nif] = alumnos.obtener_datos_alumno(nif) is not None
        if codigo not in cache_cursos:
            cache_cursos[codigo] = cursos.obtener_datos_curso(codigo) is not None

        if not cache_alumnos[nif] or not cache_cursos[codigo]:
            resumen.errores += 1
            continue

        ok = matriculas.crear_matricula(nif, codigo, fecha)

        if ok:
            resumen.nuevos += 1
        else:
            resumen.duplicados += 1

    return resumen.como_dict()